#!/usr/bin/env python3

"""
总结脚本：遍历 eval_output 文件夹中的所有结果，生成 Excel 汇总文件

本脚本是根据原来的版本改写的，主要适配当前工程中 eval_output
下文件夹命名格式，例如：

    numina-cot-dft-Qwen2.5-Math-1.5B-small_global_step_390
    numina-cot-rank_no_t-Qwen2.5-Math-1.5B-small_global_step_390
    math500-dft-Qwen2.5-Math-1.5B-small_global_step_390

解析规则：
    - 前缀支持 "numina-cot-" 或 "math500-"
    - 之后的第一段（到下一个 '-' 为止）为 method_name
    - 剩余部分（去掉 "-small_global_step_..." 的尾巴）为 model_name
"""

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


def parse_folder_name(folder_name):
    """
    从当前工程使用的文件夹名称中提取模型名称和方法名称

    例如:
        "numina-cot-dft-Qwen2.5-Math-1.5B-small_global_step_390"
            -> ("Qwen2.5-Math-1.5B", "dft")
        "numina-cot-rank_no_t-Qwen2.5-Math-1.5B-small_global_step_390"
            -> ("Qwen2.5-Math-1.5B", "rank_no_t")
        "math500-dft-Qwen2.5-Math-1.5B-small_global_step_390"
            -> ("Qwen2.5-Math-1.5B", "dft")
    """
    # 支持多种前缀
    prefixes = ["numina-cot-", "math500-"]
    prefix = None
    for p in prefixes:
        if folder_name.startswith(p):
            prefix = p
            break
    
    if prefix is None:
        return None, None

    # 去掉前缀，形如 "dft-Qwen2.5-Math-1.5B-small_global_step_390"
    rest = folder_name[len(prefix) :]

    # method_name 为第一段（到下一个 '-'）
    parts = rest.split("-", 1)
    if len(parts) < 2:
        return None, None

    method_name = parts[0]
    tail = parts[1]  # 形如 "Qwen2.5-Math-1.5B-small_global_step_390"

    # 先去掉 "_global_step..." 之类的尾巴（兼容 small / medium 等）
    # 例如：
    #   "Qwen3-8B-medium_global_step_390" -> "Qwen3-8B-medium"
    #   "Qwen2.5-Math-1.5B-small_global_step_390" -> "Qwen2.5-Math-1.5B-small"
    if "_global_step" in tail:
        model_part = tail.split("_global_step", 1)[0]
    else:
        model_part = tail

    # 再把末尾的 size_tag 去掉（如 "-small", "-medium" 等）
    # 例如：
    #   "Qwen3-8B-medium" -> "Qwen3-8B"
    #   "Qwen3-14B-small" -> "Qwen3-14B"
    size_tags = ["-small", "-medium"]
    for tag in size_tags:
        if model_part.endswith(tag):
            model_part = model_part[: -len(tag)]
            break

    model_name = model_part

    return model_name, method_name


def read_metrics_file(metrics_path):
    """
    读取 metrics JSON 文件并提取所有关键指标
    返回包含 mean_acc 和 pass_at_k 的字典
    """
    try:
        with open(metrics_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        metrics = {}

        # 提取 pass_at_k 的所有值
        pass_at_k = data.get("pass_at_k", {})
        if isinstance(pass_at_k, dict):
            for k, v in pass_at_k.items():
                metrics[k] = v

        return metrics
    except (FileNotFoundError, json.JSONDecodeError, KeyError) as e:
        print(f"Warning: Failed to read {metrics_path}: {e}")
        return {}


def collect_all_results(eval_output_dir):
    """
    收集 eval_output 下所有文件夹中的结果
    """
    results = []
    eval_output_path = Path(eval_output_dir)

    if not eval_output_path.exists():
        print(f"Error: Directory {eval_output_dir} does not exist")
        return results

    # 允许的方法列表（与原始脚本保持一致）
    allowed_methods = [
        # "ori",
        # "base",
        "overtone",
        # "dft",
        # "direct_rank_scale_no_weighting",
        # "eaft",
        # "talr",
        # "dft_rank",
        # "rev_dft_rank",
        # "rank_linear",
        # "rank_inv_linear",
        # "rank_no_t",
        # "rank_no_norm",
        # "dft_rank_no_t",
        # "dft_rank_no_norm",
        # "dft_rank_no_t_no_norm",
        # "direct_rank_scale",
        # "direct_rank_scale_no_weighting_rev",
        # "direct_rank_scale_no_prob",
        # "direct_rank_scale_no_entropy",
        # direct_rank_scale_no_weighting_log_entropy
        # "entropy_reg_002",
        # "entropy_reg_005",
        # "entropy_reg_008",
        # "alpha_power_020",
        # "alpha_power_050",
        # "alpha_power_080",
        # "direct_rank_scale_no_dft",
        # "dft_rank_no_t_no_norm_1e_5",
        # "dft_rank_no_t_no_norm_2e_5",
        # "dft_rank_no_t_no_norm_5e_5",
        # "grpo_rank_scale_290_old",
        # "grpo_rank_scale_290",
        # "grpo_vanilla_290",
        # "grpo_vanilla_200",
        # "grpo_rank_scale_200",
        # "rank_no_t_no_norm",
        # "rev_rank_no_t_no_norm",
        # "dft_rank_no_norm",
        # "dft_rank_OER_rank_no_t_no_norm",
        # "grpo_rank_scale_adv_positive_only_290",
        # "dft_rank_no_t_cut_0_95",
        # "logp_ge_02",
        # "logp_le_08",
    ]

    # 当前工程中实际使用的数据集（根据已有文件适配）
    datasets = ["aime25", "aime24", "amc23", "math_oai", "minerva_math", "olympiadbench"]

    # 遍历所有文件夹
    for folder in sorted(eval_output_path.iterdir()):
        if not folder.is_dir():
            continue

        folder_name = folder.name
        model_name, method_name = parse_folder_name(folder_name)

        if model_name is None or method_name is None:
            print(f"Warning: Skipping folder with invalid name format: {folder_name}")
            continue

        # 只处理允许的方法
        if method_name not in allowed_methods:
            continue

        # 初始化结果行
        result_row = {
            "model": model_name,
            "method": method_name,
        }

        # 定义所有可能的指标，确保所有列都被初始化
        all_metrics = ["pass@1", "pass@2", "pass@4", "pass@8", "pass@16"]

        for dataset in datasets:
            metrics_file = folder / f"{dataset}_metrics.json"

            # 先初始化所有指标列为空字符串
            for metric_name in all_metrics:
                result_row[f"{dataset}_{metric_name}"] = ""

            if metrics_file.exists():
                metrics = read_metrics_file(metrics_file)
                # 存储所有指标（覆盖初始化的空值）
                for metric_name, metric_value in metrics.items():
                    if metric_name in all_metrics:
                        result_row[f"{dataset}_{metric_name}"] = (
                            metric_value if metric_value is not None else ""
                        )
            else:
                print(f"Warning: {metrics_file} not found")

        results.append(result_row)

    return results


def write_excel(results, output_path):
    """
    将结果写入 Excel 文件，使用多行表头，并将每列的最高值加粗
    """
    if not results:
        print("No results to write")
        return

    # 数据集列表（需与 collect_all_results 中保持一致）
    datasets = ["aime25", "aime24", "amc23", "math_oai", "minerva_math", "olympiadbench"]

    # 定义每个数据集的指标顺序
    metric_names = ["pass@1", "pass@2", "pass@4", "pass@8", "pass@16"]

    # 构建列结构：每个数据集对应多个指标列
    column_structure = []
    # 前两列是 model 和 method
    column_structure.append(("model", "model"))
    column_structure.append(("method", "method"))

    # 为每个数据集添加所有指标列
    for dataset in datasets:
        for metric in metric_names:
            column_structure.append((dataset, metric))

    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # 写入多行表头
    # 第一行：数据集名称（model 和 method 列合并）
    bold_font = Font(bold=True)
    col_idx = 1

    # model 列
    cell = ws.cell(row=1, column=col_idx)
    cell.value = "model"
    cell.font = bold_font
    ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
    col_idx += 1

    # method 列
    cell = ws.cell(row=1, column=col_idx)
    cell.value = "method"
    cell.font = bold_font
    ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
    col_idx += 1

    # 为每个数据集写入表头
    for dataset in datasets:
        # 第一行：数据集名称（合并该数据集的所有列）
        start_col = col_idx
        end_col = col_idx + len(metric_names) - 1
        cell = ws.cell(row=1, column=col_idx)
        cell.value = dataset
        cell.font = bold_font
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

        # 第二行：指标名称
        for metric in metric_names:
            cell = ws.cell(row=2, column=col_idx)
            cell.value = metric
            cell.font = bold_font
            col_idx += 1

    # 写入数据并找出每个模型在每个列的最高值
    numeric_columns = {}  # 存储数值列的索引和值（包含模型信息）
    numeric_col_indices = []  # 数值列的索引列表

    # 确定数值列的索引（除了 model 和 method 列）
    for col_idx, (dataset, metric) in enumerate(column_structure, start=1):
        if dataset not in ["model", "method"]:
            numeric_col_indices.append(col_idx)
            numeric_columns[col_idx] = []

    # 写入数据行（从第3行开始，因为前两行是表头）
    for row_idx, result in enumerate(results, start=3):
        model_name = result.get("model", "")
        col_idx = 1

        # 写入 model 和 method
        for col_name in ["model", "method"]:
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = result.get(col_name, "")
            col_idx += 1

        # 写入每个数据集的每个指标
        for dataset in datasets:
            for metric in metric_names:
                cell = ws.cell(row=row_idx, column=col_idx)
                fieldname = f"{dataset}_{metric}"
                value = result.get(fieldname, "")

                # 如果是数值列，尝试转换为浮点数
                if col_idx in numeric_col_indices:
                    try:
                        if value != "":
                            num_value = float(value)
                            cell.value = num_value
                            # 存储 (行号, 值, 模型名) 元组
                            numeric_columns[col_idx].append((row_idx, num_value, model_name))
                        else:
                            cell.value = ""
                    except (ValueError, TypeError):
                        cell.value = value
                else:
                    cell.value = value

                col_idx += 1

    # 找出每个模型在每个列的最高值并加粗
    bold_font = Font(bold=True)
    for col_idx in numeric_col_indices:
        if numeric_columns[col_idx]:
            # 按模型分组
            model_groups = {}
            for row_idx, num_value, model_name in numeric_columns[col_idx]:
                if model_name not in model_groups:
                    model_groups[model_name] = []
                model_groups[model_name].append((row_idx, num_value))

            # 对每个模型组，找出最高值并加粗
            for model_name, group_data in model_groups.items():
                if group_data:
                    # 找出该模型组内的最高值
                    max_row, max_value = max(group_data, key=lambda x: x[1])

                    # 找出该模型组内所有等于最高值的行（处理并列情况）
                    max_rows = [row for row, val in group_data if val == max_value]

                    # 将这些单元格加粗
                    for row in max_rows:
                        cell = ws.cell(row=row, column=col_idx)
                        cell.font = bold_font

    # 保存文件
    wb.save(output_path)

    print(f"Summary written to {output_path}")
    print(f"Total rows: {len(results)}")


def main():
    """
    主函数
    """
    parser = argparse.ArgumentParser(description="总结脚本：遍历 eval_output 文件夹中的所有结果，生成 Excel 汇总文件")
    parser.add_argument(
        "--eval-output-dir",
        type=str,
        default="eval_output_5e-5",
        help="指定 eval_output 目录路径（默认：脚本目录下的 eval_output）"
    )
    args = parser.parse_args()

    # 获取当前脚本所在目录
    script_dir = Path(__file__).parent
    
    # 如果指定了参数，使用指定的路径；否则使用默认路径
    if args.eval_output_dir:
        eval_output_dir = script_dir / args.eval_output_dir
    else:
        eval_output_dir = script_dir / "eval_output"
    
    output_excel = eval_output_dir / "summary_results_numina_cot.xlsx"

    print(f"Scanning directory: {eval_output_dir}")

    # 收集所有结果
    results = collect_all_results(eval_output_dir)

    # 为了让相同 model 的结果靠在一起显示，这里按照 model 名称排序
    # 若某些结果中缺少 model 字段，则使用空字符串作为默认值
    results = sorted(results, key=lambda x: x.get("model", ""))

    if results:
        # 写入 Excel
        write_excel(results, output_excel)
        print("\nSummary completed successfully!")
    else:
        print("No results found. Please check the eval_output directory.")


if __name__ == "__main__":
    main()


