#!/usr/bin/env python3
"""Evaluate CEN on SAD dataset using official repo code."""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

import numpy as np

sys.path.insert(0, "/repo/src")
from autofocus_cen.io import stream_blocks_from_dat
from autofocus_cen.cen import CENConfig, estimate_focus_from_blocks


# Paper Table 8 scenes (15 total, excluding USAF_static_significant)
PAPER_SCENES = {
    "USAF_static_constant", "USAF_static_decrease", "USAF_static_increase",
    "USAF_static_minor",
    "USAF_dynamic_constant", "USAF_dynamic_decrease", "USAF_dynamic_increase",
    "USAF_dynamic_minor", "USAF_dynamic_significant",
    "Bottle_static_constant", "Bottle_dynamic_constant",
    "Lily_static_constant", "Lily_dynamic_constant",
    "Fan_static_constant", "Fan_dynamic_constant",
}

TARGET_PREFIXES = ["Bottle", "Lily", "Fan", "USAF"]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--sad_root", default="/datasets/sad_dataset")
    parser.add_argument("--output", default="./results_sad_cen.csv")
    args = parser.parse_args()

    sad_root = Path(args.sad_root)
    config_file = sad_root / "config.xlsx"

    import openpyxl
    wb = openpyxl.load_workbook(config_file)
    ws = wb.active
    headers = [str(cell.value) for cell in ws[1]]
    print(f"Config: {headers}")

    scenes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        s = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        scenes.append(s)

    # Filter to paper scenes
    target_scenes = [s for s in scenes if s["Scene Name"] in PAPER_SCENES]
    print(f"Paper scenes: {len(target_scenes)}")

    # Also filter to only Bottle/Lily/Fan/USAF scenes
    target_scenes_final = []
    for s in target_scenes:
        name = str(s["Scene Name"])
        for p in TARGET_PREFIXES:
            if name.lower().startswith(p.lower()):
                target_scenes_final.append(s)
                break
    target_scenes = target_scenes_final
    print(f"Filtered to {', '.join(TARGET_PREFIXES)}: {len(target_scenes)}")

    config = CENConfig()
    results = []
    scene_groups = {}

    for s in target_scenes:
        name = str(s["Scene Name"])
        dt = int(s["dt"])
        focal_frame = int(s["Focal frame"])
        gt_block = focal_frame // dt
        filepath = str(s["filepath"])

        dat_path = sad_root / filepath
        if not dat_path.exists():
            if filepath.startswith("./datasets/"):
                dat_path = sad_root / filepath[len("./datasets/"):]
        if not dat_path.exists():
            print(f"SKIP {name}: file not found: {filepath}")
            continue

        print(f"\n--- {name} ---")
        print(f"  dt={dt} gt_block={gt_block} (focal_frame={focal_frame})")

        try:
            blocks = [b for _, b in stream_blocks_from_dat(str(dat_path), dt=dt)]
            result = estimate_focus_from_blocks(blocks, config=config)
        except Exception as e:
            print(f"  ERROR: {e}")
            import traceback; traceback.print_exc()
            continue

        pred_block = int(result.focus_block)
        abs_err = abs(pred_block - gt_block)
        num_blocks = int(result.block_ids.size)
        rel_err_pct = (abs_err / max(num_blocks, 1)) * 100.0

        row = {
            "scene": name, "dt": dt, "gt_block": gt_block,
            "pred_block": pred_block, "abs_err": abs_err,
            "rel_err_pct": round(rel_err_pct, 2),
            "num_blocks": num_blocks, "r2": float(result.r2),
        }
        results.append(row)
        print(f"  pred={pred_block} gt={gt_block} abs_err={abs_err} "
              f"rel_err={rel_err_pct:.2f}% r2={result.r2:.4f} num_blocks={num_blocks}")

        for p in TARGET_PREFIXES:
            if name.lower().startswith(p.lower()):
                scene_groups.setdefault(p, []).append(row)
                break

    # Aggregates
    print("\n" + "=" * 60)
    print("RESULTS (15 paper scenes)")
    print("=" * 60)
    output_rows = list(results)

    for prefix in TARGET_PREFIXES:
        group = scene_groups.get(prefix, [])
        if group:
            mean_abs = np.mean([r["abs_err"] for r in group])
            mean_rel = np.mean([r["rel_err_pct"] for r in group])
            print(f"  {prefix}: abs_err={mean_abs:.2f} rel_err={mean_rel:.2f}% "
                  f"({len(group)} sub-scenes)")
            output_rows.append({
                "scene": f"{prefix}_MEAN", "dt": "", "gt_block": "",
                "pred_block": "", "abs_err": round(mean_abs, 2),
                "rel_err_pct": round(mean_rel, 2), "num_blocks": "", "r2": "",
            })

    if results:
        overall_abs = np.mean([r["abs_err"] for r in results])
        overall_rel = np.mean([r["rel_err_pct"] for r in results])
        print(f"  MEAN: abs_err={overall_abs:.2f} rel_err={overall_rel:.2f}% "
              f"({len(results)} scenes)")
        output_rows.append({
            "scene": "MEAN", "dt": "", "gt_block": "", "pred_block": "",
            "abs_err": round(overall_abs, 2),
            "rel_err_pct": round(overall_rel, 2),
            "num_blocks": "", "r2": "",
        })

    output_path = Path(args.output)
    fieldnames = ["scene", "dt", "gt_block", "pred_block", "abs_err",
                  "rel_err_pct", "num_blocks", "r2"]
    with output_path.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(output_rows)
    print(f"\nWrote: {output_path}")


if __name__ == "__main__":
    main()
