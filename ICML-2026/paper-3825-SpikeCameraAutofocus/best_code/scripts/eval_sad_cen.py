#!/usr/bin/env python3
"""Evaluate CEN on SAD dataset scenes (Bottle, Lily, Fan, USAF)."""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

import numpy as np

try:
    from autofocus_cen import CENConfig, estimate_focus_from_blocks
except ModuleNotFoundError:
    sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
    from autofocus_cen import CENConfig, estimate_focus_from_blocks

SAD_WIDTH = 400
SAD_HEIGHT = 250
SAD_PADDING_BITS = 16


def stream_blocks_from_dat(path, dt, width=SAD_WIDTH, height=SAD_HEIGHT,
                           padding_bits=SAD_PADDING_BITS):
    if dt <= 0:
        raise ValueError("dt must be positive")
    bytes_per_row = (width + padding_bits) // 8
    bytes_per_frame = bytes_per_row * height
    data = Path(path).read_bytes()
    total_frames = len(data) // bytes_per_frame

    offset = 0
    block_sum = np.zeros((height, width), dtype=np.float32)
    count = 0
    block_id = 0

    for _ in range(total_frames):
        frame_bytes = np.frombuffer(
            data[offset : offset + bytes_per_frame],
            dtype=np.uint8,
        ).reshape((bytes_per_row, height), order="F")
        offset += bytes_per_frame

        useful = frame_bytes[: width // 8, :]
        flat_bytes = useful.reshape(-1, order="F")
        bits = np.unpackbits(flat_bytes).reshape(-1, 8)[:, ::-1].reshape(-1)
        frame = bits.reshape((width, height), order="F").T.astype(np.float32)

        block_sum += frame
        count += 1
        if count == dt:
            yield block_id, block_sum.copy()
            block_id += 1
            block_sum.fill(0.0)
            count = 0


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--sad_root", default="/datasets/sad_dataset")
    parser.add_argument("--output", default="./results_sad_cen.csv")
    args = parser.parse_args()

    sad_root = Path(args.sad_root)
    config_file = sad_root / "config.xlsx"

    import openpyxl
    wb = openpyxl.load_workbook(config_file)
    ws = wb.active
    headers = [str(cell.value) for cell in ws[1]]
    print(f"Config: {headers}")

    # Parse scenes
    scenes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        s = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        scenes.append(s)

    # Filter for target scenes
    target_prefixes = ["Bottle", "Lily", "Fan", "USAF"]
    target_scenes = []
    for s in scenes:
        name = str(s["Scene Name"])
        for p in target_prefixes:
            if name.lower().startswith(p.lower()):
                target_scenes.append(s)
                break

    print(f"Target scenes: {len(target_scenes)}")

    config = CENConfig()
    results = []
    scene_groups = {}

    for s in target_scenes:
        name = str(s["Scene Name"])
        dt = int(s["dt"])
        focal_frame = int(s["Focal frame"])
        gt_block = focal_frame // dt
        filepath = str(s["filepath"])

        # Resolve path
        dat_path = sad_root / filepath
        if not dat_path.exists():
            # Strip ./datasets/ prefix if present
            if filepath.startswith("./datasets/"):
                dat_path = sad_root / filepath[len("./datasets/"):]
        if not dat_path.exists():
            print(f"SKIP {name}: file not found: {filepath}")
            continue

        print(f"\n--- {name} ---")
        print(f"  dt={dt} gt_block={gt_block} (focal_frame={focal_frame}) file={dat_path}")

        try:
            blocks_iter = (block for _, block in stream_blocks_from_dat(dat_path, dt=dt))
            result = estimate_focus_from_blocks(blocks_iter, config=config)
        except Exception as e:
            print(f"  ERROR: {e}")
            import traceback; traceback.print_exc()
            continue

        pred_block = int(result.focus_block)
        abs_err = abs(pred_block - gt_block)
        num_blocks = int(result.block_ids.size)
        rel_err_pct = (abs_err / max(num_blocks, 1)) * 100.0

        row = {
            "scene": name, "dt": dt, "gt_block": gt_block,
            "pred_block": pred_block, "abs_err": abs_err,
            "rel_err_pct": round(rel_err_pct, 2),
            "num_blocks": num_blocks, "r2": float(result.r2),
        }
        results.append(row)
        print(f"  pred={pred_block} gt={gt_block} abs_err={abs_err} "
              f"rel_err={rel_err_pct:.2f}% r2={result.r2:.4f}")

        for p in target_prefixes:
            if name.lower().startswith(p.lower()):
                scene_groups.setdefault(p, []).append(row)
                break

    # Aggregates
    print("\n" + "=" * 60)
    print("RESULTS")
    print("=" * 60)
    output_rows = list(results)

    for prefix in target_prefixes:
        group = scene_groups.get(prefix, [])
        if group:
            mean_abs = np.mean([r["abs_err"] for r in group])
            mean_rel = np.mean([r["rel_err_pct"] for r in group])
            print(f"  {prefix}: abs_err={mean_abs:.2f} rel_err={mean_rel:.2f}% "
                  f"({len(group)} sub-scenes)")
            output_rows.append({
                "scene": f"{prefix}_MEAN", "dt": "", "gt_block": "",
                "pred_block": "", "abs_err": round(mean_abs, 2),
                "rel_err_pct": round(mean_rel, 2), "num_blocks": "", "r2": "",
            })

    if results:
        overall_abs = np.mean([r["abs_err"] for r in results])
        overall_rel = np.mean([r["rel_err_pct"] for r in results])
        print(f"  MEAN: abs_err={overall_abs:.2f} rel_err={overall_rel:.2f}% "
              f"({len(results)} scenes)")
        output_rows.append({
            "scene": "MEAN", "dt": "", "gt_block": "", "pred_block": "",
            "abs_err": round(overall_abs, 2),
            "rel_err_pct": round(overall_rel, 2),
            "num_blocks": "", "r2": "",
        })

    # CSV
    output_path = Path(args.output)
    fieldnames = ["scene", "dt", "gt_block", "pred_block", "abs_err",
                  "rel_err_pct", "num_blocks", "r2"]
    with output_path.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(output_rows)
    print(f"\nWrote: {output_path}")


if __name__ == "__main__":
    main()
